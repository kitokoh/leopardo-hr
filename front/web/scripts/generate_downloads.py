#!/usr/bin/env python3
"""
Generate the guide/checklist/template files served under
front/web/public/downloads/.

These are referenced by /guides/rh-startup, /guides/checklist-paie and
/guides/planning-employes as free lead-gen downloads, and by the
whitelist in app/api/downloads/route.ts. Content is derived from the
existing blog posts (front/web/src/content/blog/*) that cover the same
topics, so this script can be re-run whenever that source content
changes.

Usage: python3 front/web/scripts/generate_downloads.py
"""
import os

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "..", "public", "downloads")

EMERALD = colors.HexColor("#059669")
SLATE = colors.HexColor("#1e293b")


def build_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="LeopardoTitle", fontSize=24, leading=28, spaceAfter=6,
        textColor=SLATE, fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        name="LeopardoSubtitle", fontSize=12, leading=16, spaceAfter=18,
        textColor=colors.HexColor("#475569"),
    ))
    styles.add(ParagraphStyle(
        name="LeopardoH2", fontSize=15, leading=19, spaceBefore=16,
        spaceAfter=8, textColor=EMERALD, fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        name="LeopardoH3", fontSize=12, leading=15, spaceBefore=8,
        spaceAfter=4, textColor=SLATE, fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        name="LeopardoBody", fontSize=10.5, leading=15, spaceAfter=6,
        textColor=colors.HexColor("#334155"),
    ))
    styles.add(ParagraphStyle(
        name="LeopardoFooter", fontSize=9, leading=13,
        textColor=colors.HexColor("#64748b"), spaceBefore=20,
    ))
    return styles


def guide_rh_startup_pdf(path):
    styles = build_styles()
    doc = SimpleDocTemplate(
        path, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title="Guide Complet RH pour Startup", author="Leopardo RH",
    )
    story = [
        Paragraph("🐆 Leopardo RH", styles["LeopardoSubtitle"]),
        Paragraph("Guide Complet RH pour Startup", styles["LeopardoTitle"]),
        Paragraph(
            "Tout ce que vous devez savoir pour gérer vos employés en startup",
            styles["LeopardoSubtitle"],
        ),
        Paragraph(
            "Gérer les ressources humaines dans une startup est un défi unique. "
            "Vous devez être agile, efficace et créatif avec des ressources "
            "limitées. Ce guide vous aidera à mettre en place une gestion RH "
            "solide dès le départ.",
            styles["LeopardoBody"],
        ),
    ]

    chapters = [
        ("1. Les Fondamentaux de la Gestion RH en Startup", [
            "Pourquoi la RH est importante : attirer et retenir les meilleurs "
            "talents, maintenir une culture forte, assurer la conformité "
            "légale, optimiser la productivité.",
            "Les 3 piliers de la RH startup : Recrutement, Développement, "
            "Rétention.",
        ]),
        ("2. Recrutement et Onboarding", [
            "Définissez vos besoins avant de recruter : rôle, compétences "
            "essentielles, budget.",
            "Processus efficace : définir le profil, sourcer les candidats, "
            "entretiens, vérification des références, offre compétitive.",
            "Un bon onboarding réduit le turnover de 50 % : accueil "
            "personnalisé, formation aux outils, présentation de l'équipe, "
            "clarification des attentes, mentor ou buddy.",
        ]),
        ("3. Gestion des Contrats et Conformité", [
            "Types de contrats : CDI, CDD, Stage, Freelance.",
            "Conformité légale : droit du travail, mise à jour des contrats, "
            "gestion des congés, cotisations sociales, dossiers employés.",
        ]),
        ("4. Gestion de la Paie", [
            "Éléments de la paie : salaire brut, cotisations sociales, impôt "
            "sur le revenu, avantages (tickets restaurant, mutuelle, etc.).",
            "Bonnes pratiques : payer à temps (avant le 5 du mois), utiliser "
            "un logiciel de paie, conserver les bulletins, communiquer "
            "clairement.",
        ]),
        ("5. Gestion des Absences et Congés", [
            "Types de congés : congés payés (25 jours minimum en France), "
            "maladie, maternité/paternité, jours fériés.",
            "Gestion efficace : calendrier partagé, approbation rapide, suivi "
            "des soldes, communication claire des politiques.",
        ]),
        ("6. Culture et Engagement", [
            "Créer une bonne culture : définir vos valeurs, communiquer "
            "régulièrement, célébrer les succès, écouter vos employés, "
            "offrir du développement.",
            "Engagement : réunions 1-on-1 régulières, feedback constructif, "
            "opportunités de croissance, reconnaissance, équilibre "
            "travail-vie.",
        ]),
        ("7. Outils et Systèmes", [
            "Outils essentiels : logiciel RH, paie, pointage, gestion des "
            "documents, communication (Slack, Teams, etc.).",
            "Mise en place progressive : commencez simple, automatisez "
            "progressivement, investissez quand vous grandissez, choisissez "
            "des outils intégrés.",
        ]),
        ("8. Gestion des Performances", [
            "Évaluations régulières : annuelles, feedback continu, objectifs "
            "clairs, plans de développement.",
            "Gestion des problèmes : documenter, communiquer clairement, "
            "donner des chances d'amélioration, suivre les progrès.",
        ]),
        ("9. Santé et Sécurité", [
            "Responsabilités légales : évaluation des risques, formation à "
            "la sécurité, équipement de protection, signalement des "
            "accidents, assurance.",
            "Bien-être : environnement de travail sain, prévention du "
            "burnout, soutien mental, team building.",
        ]),
        ("10. Croissance et Scalabilité", [
            "Préparer la croissance : documenter les processus, créer des "
            "politiques claires, former les managers, investir dans les "
            "outils, planifier les embauches.",
            "Transition vers une équipe RH dédiée : embauchez un responsable "
            "RH, déléguez l'administratif, formalisez les processus, "
            "continuez à cultiver la culture.",
        ]),
    ]

    for title, points in chapters:
        story.append(Paragraph(title, styles["LeopardoH2"]))
        story.append(ListFlowable(
            [ListItem(Paragraph(p, styles["LeopardoBody"])) for p in points],
            bulletType="bullet", start="•",
        ))

    story.append(Paragraph(
        "La gestion RH en startup ne doit pas être compliquée. Commencez "
        "par les bases, automatisez progressivement, et adaptez-vous à "
        "votre croissance. L'important est de traiter vos employés avec "
        "respect et de créer un environnement où ils peuvent s'épanouir.",
        styles["LeopardoH2"],
    ))
    story.append(Paragraph(
        "Besoin d'aide ? Leopardo RH simplifie votre gestion RH avec une "
        "plateforme complète et facile à utiliser — recrutement, paie, "
        "absences, pointage et plus, en un seul endroit. "
        "https://leopardo.app",
        styles["LeopardoFooter"],
    ))

    doc.build(story)


def checklist_paie_pdf(path):
    styles = build_styles()
    doc = SimpleDocTemplate(
        path, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title="Checklist Paie 2024", author="Leopardo RH",
    )
    story = [
        Paragraph("🐆 Leopardo RH", styles["LeopardoSubtitle"]),
        Paragraph("Checklist Paie 2024", styles["LeopardoTitle"]),
        Paragraph(
            "Assurez la conformité de votre paie 2024 avec cette checklist "
            "complète.",
            styles["LeopardoSubtitle"],
        ),
    ]

    sections = [
        ("Avant la Paie", [
            "Vérifiez les données des employés",
            "Mettez à jour les salaires",
            "Vérifiez les cotisations",
            "Validez les absences",
            "Vérifiez les avantages",
            "Testez les calculs",
            "Préparez les documents",
        ]),
        ("Pendant la Paie", [
            "Calculez les salaires bruts",
            "Appliquez les cotisations",
            "Calculez l'impôt",
            "Appliquez les déductions",
            "Générez les bulletins",
            "Vérifiez les totaux",
            "Validez les anomalies",
        ]),
        ("Après la Paie", [
            "Vérifiez les bulletins",
            "Comparez avec la paie précédente",
            "Validez les exports comptables",
            "Archivez les documents",
            "Envoyez les bulletins",
            "Collectez les retours",
            "Documentez les changements",
        ]),
        ("Conformité", [
            "Respectez les taux 2024",
            "Mettez à jour les cotisations",
            "Vérifiez les seuils",
            "Appliquez les augmentations",
            "Respectez les délais",
            "Conservez les documents",
            "Auditez régulièrement",
        ]),
        ("Sécurité", [
            "Chiffrez les données",
            "Contrôlez l'accès",
            "Sauvegardez régulièrement",
            "Testez les restaurations",
            "Documentez les accès",
            "Formez votre équipe",
            "Auditez la sécurité",
        ]),
        ("Outils", [
            "Logiciel de paie à jour",
            "Intégration comptable",
            "Rapports automatiques",
            "Alertes de conformité",
            "Archivage sécurisé",
            "Support disponible",
        ]),
    ]

    for title, items in sections:
        story.append(Paragraph(title, styles["LeopardoH2"]))
        story.append(ListFlowable(
            [ListItem(Paragraph(f"☐ {i}", styles["LeopardoBody"]))
             for i in items],
            bulletType="bullet", start="",
        ))

    story.append(Paragraph(
        "Besoin d'aide ? Leopardo RH automatise votre paie et garantit la "
        "conformité 2024. https://leopardo.app",
        styles["LeopardoFooter"],
    ))

    doc.build(story)


def planning_employes_xlsx(path):
    wb = Workbook()

    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(*(Side(style="thin", color="CBD5E1"),) * 4)
    center = Alignment(horizontal="center", vertical="center")

    def style_header_row(ws, row, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

    # --- Sheet 1: Employés ---
    ws = wb.active
    ws.title = "Employés"
    headers = ["Nom", "Poste", "Département", "Disponibilité", "Préférences"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    sample_rows = [
        ["Jean Dupont", "Développeur", "Tech", "Temps plein", "Matin"],
        ["Marie Curie", "RH Manager", "RH", "Temps plein", "—"],
        ["Ali Ben Salem", "Support Client", "Support", "Mi-temps", "Après-midi"],
    ]
    for r in sample_rows:
        ws.append(r)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # --- Sheet 2: Planning Mensuel ---
    ws2 = wb.create_sheet("Planning Mensuel")
    days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    ws2.append(["Employé"] + days)
    style_header_row(ws2, 1, len(days) + 1)
    for r in sample_rows:
        ws2.append([r[0]] + ["" for _ in days])
    ws2.column_dimensions["A"].width = 20
    for col in range(2, len(days) + 2):
        ws2.column_dimensions[get_column_letter(col)].width = 14
    ws2.append([])
    ws2.append(["Légende : T = Travail, C = Congé, M = Maladie, — = Repos"])

    # --- Sheet 3: Heures de Travail ---
    ws3 = wb.create_sheet("Heures de Travail")
    headers3 = ["Employé", "Heures prévues", "Heures réelles", "Pauses (min)", "Écart"]
    ws3.append(headers3)
    style_header_row(ws3, 1, len(headers3))
    for r in sample_rows:
        ws3.append([r[0], 35, 35, 30, 0])
    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 20

    # --- Sheet 4: Rapports ---
    ws4 = wb.create_sheet("Rapports")
    ws4.append(["Indicateur", "Valeur"])
    style_header_row(ws4, 1, 2)
    ws4.append(["Total heures planifiées", "=SUM('Heures de Travail'!B2:B100)"])
    ws4.append(["Total heures réelles", "=SUM('Heures de Travail'!C2:C100)"])
    ws4.append(["Nombre d'employés", "=COUNTA(Employés!A2:A100)"])
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 25

    wb.save(path)


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    guide_rh_startup_pdf(os.path.join(OUT_DIR, "guide-rh-startup.pdf"))
    checklist_paie_pdf(os.path.join(OUT_DIR, "checklist-paie-2024.pdf"))
    planning_employes_xlsx(os.path.join(OUT_DIR, "modele-planning-employes.xlsx"))
    print("Generated:")
    for f in ["guide-rh-startup.pdf", "checklist-paie-2024.pdf", "modele-planning-employes.xlsx"]:
        p = os.path.join(OUT_DIR, f)
        print(f"  {p} ({os.path.getsize(p)} bytes)")


if __name__ == "__main__":
    main()
